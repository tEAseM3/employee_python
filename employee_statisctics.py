import csv
from datetime import datetime
from openpyxl import load_workbook
import matplotlib.pyplot as plt

def calculate_age(birthdate):
    """Обчислює вік на поточну дату."""
    today = datetime.today()
    age = today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))
    return age

def count_gender_from_csv(csv_filename):
    """Рахує кількість співробітників за статтю з CSV-файлу та будує відповідну діаграму."""
    gender_counts = {"Чоловік": 0, "Жінка": 0}
    
    try:
        with open(csv_filename, newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            
            for row in reader:
                gender = row["Стать"]
                
                # Збільшуємо лічильник в залежності від статі
                if gender == "Чоловік":
                    gender_counts["Чоловік"] += 1
                elif gender == "Жінка":
                    gender_counts["Жінка"] += 1

        # Виводимо результати в консоль
        print("Кількість співробітників за статтю:")
        for gender, count in gender_counts.items():
            print(f"{gender}: {count}")
        
        # Створюємо діаграму
        plt.figure(figsize=(6, 6))
        plt.pie(
            gender_counts.values(),
            labels=gender_counts.keys(),
            autopct='%1.1f%%',
            startangle=140
        )
        plt.title("Кількість співробітників за статтю")
        plt.show()
    
    except FileNotFoundError:
        print("Помилка: файл CSV не знайдено.")
    except Exception as e:
        print(f"Сталася помилка: {e}")

def count_by_age_category(xlsx_filename):
    """Рахує кількість співробітників у кожній віковій категорії та будує діаграму."""
    try:
        # Відкриваємо файл XLSX
        wb = load_workbook(xlsx_filename)
        ws = wb['all']
        
        # Ініціалізуємо лічильники для вікових категорій
        age_categories = {
            "до 18": 0,
            "18-45": 0,
            "45-70": 0,
            "понад 70": 0
        }
        
        # Перебираємо рядки з даними
        for row in ws.iter_rows(min_row=2, values_only=True):  # min_row=2 пропускає заголовок
            age = row[5]  # Вік у шостому стовпці
            
            # Збільшуємо лічильники в залежності від вікової категорії
            if age < 18:
                age_categories["до 18"] += 1
            elif 18 <= age <= 45:
                age_categories["18-45"] += 1
            elif 45 < age <= 70:
                age_categories["45-70"] += 1
            else:
                age_categories["понад 70"] += 1
        
        # Виводимо результати в консоль
        print("Кількість співробітників у кожній віковій категорії:")
        for category, count in age_categories.items():
            print(f"{category}: {count}")
        
        # Створюємо діаграму
        plt.figure(figsize=(8, 6))
        plt.bar(age_categories.keys(), age_categories.values(), color=['blue', 'green', 'orange', 'red'])
        plt.xlabel('Вікова категорія')
        plt.ylabel('Кількість співробітників')
        plt.title('Кількість співробітників у кожній віковій категорії')
        plt.show()
        
    except FileNotFoundError:
        print("Помилка: файл XLSX не знайдено.")
    except Exception as e:
        print(f"Сталася помилка: {e}")

def count_gender_by_age_category(xlsx_filename):
    """Рахує кількість співробітників чоловічої і жіночої статі в кожній віковій категорії та будує діаграми."""
    try:
        # Відкриваємо файл XLSX
        wb = load_workbook(xlsx_filename)
        ws = wb['all']
        
        # Ініціалізуємо лічильники для кожної категорії та статі
        gender_age_categories = {
            "до 18": {"Чоловік": 0, "Жінка": 0},
            "18-45": {"Чоловік": 0, "Жінка": 0},
            "45-70": {"Чоловік": 0, "Жінка": 0},
            "понад 70": {"Чоловік": 0, "Жінка": 0}
        }
        
        # Перебираємо рядки з даними
        for row in ws.iter_rows(min_row=2, values_only=True):
            age = row[5]  # Вік у шостому стовпці
            gender = row[6]  # Стать у сьомому стовпці
            
            # Визначаємо категорію віку
            if age < 18:
                category = "до 18"
            elif 18 <= age <= 45:
                category = "18-45"
            elif 45 < age <= 70:
                category = "45-70"
            else:
                category = "понад 70"
            
            # Збільшуємо лічильники в залежності від статі
            if gender == "Чоловік":
                gender_age_categories[category]["Чоловік"] += 1
            elif gender == "Жінка":
                gender_age_categories[category]["Жінка"] += 1

        # Виводимо результати в консоль
        print("Кількість співробітників чоловічої і жіночої статі в кожній віковій категорії:")
        for category, counts in gender_age_categories.items():
            print(f"{category}: Чоловіки - {counts['Чоловік']}, Жінки - {counts['Жінка']}")
        
        # Створюємо діаграми
        for category, counts in gender_age_categories.items():
            plt.figure(figsize=(6, 6))
            plt.pie(
                [counts['Чоловік'], counts['Жінка']],
                labels=["Чоловіки", "Жінки"],
                autopct='%1.1f%%',
                startangle=140
            )
            plt.title(f"Стать співробітників у категорії {category}")
            plt.show()
        
    except FileNotFoundError:
        print("Помилка: файл XLSX не знайдено.")
    except Exception as e:
        print(f"Сталася помилка: {e}")

# Запуск функцій
csv_filename = 'employees.csv'
xlsx_filename = 'employees_categories.xlsx'

# Побудова діаграми за статтю на основі даних з CSV
count_gender_from_csv(csv_filename)

# Побудова діаграм за віковими категоріями на основі даних з XLSX
count_by_age_category(xlsx_filename)
count_gender_by_age_category(xlsx_filename)
