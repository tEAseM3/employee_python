import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException

def calculate_age(birthdate):
    """Обчислює вік на поточну дату."""
    today = datetime.today()
    age = today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))
    return age

def create_xlsx_from_csv(csv_filename, xlsx_filename):
    try:
        # Читання даних з CSV файлу
        employees = []
        with open(csv_filename, newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                # Перетворюємо рядок з датою народження на об'єкт datetime
                try:
                    birthdate = datetime.strptime(row["Дата народження"], '%d.%m.%Y')
                except ValueError:
                    print(f"Помилка при перетворенні дати народження: {row['Дата народження']}")
                    continue
                
                # Обчислюємо вік
                age = calculate_age(birthdate)
                
                # Додаємо запис з віком та статтю
                employees.append({
                    "№": len(employees) + 1,
                    "Прізвище": row["Прізвище"],
                    "Ім'я": row["Ім'я"],
                    "По батькові": row["По батькові"],
                    "Дата народження": row["Дата народження"],
                    "Вік": age,
                    "Стать": row["Стать"]
                })
                
        # Створюємо новий файл XLSX
        wb = Workbook()
        
        # Створюємо листи
        ws_all = wb.active
        ws_all.title = "all"
        ws_younger_18 = wb.create_sheet("younger_18")
        ws_18_45 = wb.create_sheet("18-45")
        ws_45_70 = wb.create_sheet("45-70")
        ws_older_70 = wb.create_sheet("older_70")
        
        # Заголовки для таблиць
        headers = ["№", "Прізвище", "Ім'я", "По батькові", "Дата народження", "Вік", "Стать"]
        
        # Заповнюємо кожен лист заголовками
        for ws in [ws_all, ws_younger_18, ws_18_45, ws_45_70, ws_older_70]:
            ws.append(headers)
        
        # Записуємо дані на відповідні листи
        for employee in employees:
            ws_all.append(list(employee.values()))
            
            if employee["Вік"] < 18:
                ws_younger_18.append(list(employee.values()))
            elif 18 <= employee["Вік"] <= 45:
                ws_18_45.append(list(employee.values()))
            elif 45 < employee["Вік"] <= 70:
                ws_45_70.append(list(employee.values()))
            else:
                ws_older_70.append(list(employee.values()))
        
        # Зберігаємо файл XLSX
        wb.save(xlsx_filename)
        print("Програма завершила свою роботу успішно.")
    
    except FileNotFoundError:
        print("Помилка: файл CSV не знайдено.")
    except InvalidFileException:
        print("Помилка при створенні XLSX файлу.")
    except Exception as e:
        print(f"Сталася непередбачувана помилка: {e}")

# Запуск програми
create_xlsx_from_csv('employees.csv', 'employees_categories.xlsx')
